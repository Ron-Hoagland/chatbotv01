
import os
from pdfminer.high_level import extract_text as extract_pdf_text
from docx import Document
from openpyxl import load_workbook
from langchain_text_splitters import RecursiveCharacterTextSplitter
from sentence_transformers import SentenceTransformer
import numpy as np
import faiss
def embed_chunks(chunks, model_name="all-MiniLM-L6-v2"):
	"""Embed text chunks using a local sentence-transformers model."""
	model = SentenceTransformer(model_name)
	embeddings = model.encode(chunks, show_progress_bar=True)
	return embeddings

def build_vector_store(chunks, model_name="all-MiniLM-L6-v2"):
	"""Build a FAISS vector store from text chunks and return the store object."""
	embeddings = embed_chunks(chunks, model_name=model_name)
	dim = embeddings.shape[1] if hasattr(embeddings, 'shape') else len(embeddings[0])
	index = faiss.IndexFlatL2(dim)
	index.add(np.array(embeddings))
	return {"index": index, "chunks": chunks, "embeddings": embeddings, "model_name": model_name}

def retrieve_relevant_chunks(vector_store, question, k=5):
	"""Query the vector store for the top-k relevant chunks to the user's question using local embeddings."""
	model = SentenceTransformer(vector_store["model_name"])
	question_emb = model.encode([question])
	D, I = vector_store["index"].search(np.array(question_emb), k)
	return [vector_store["chunks"][i] for i in I[0]]

import requests

def generate_answer(question, relevant_chunks, ollama_url="http://localhost:11434/api/generate", model="llama3"):
	"""Generate a context-aware answer using a local Ollama LLM."""
	context = '\n'.join(relevant_chunks)
	prompt = f"You are a helpful assistant. Use the following context to answer the user's question.\n\nContext:\n{context}\n\nQuestion: {question}\nAnswer:"
	payload = {
		"model": model,
		"prompt": prompt,
		"stream": False
	}
	try:
		response = requests.post(ollama_url, json=payload)
		response.raise_for_status()
		return response.json().get("response", "No answer generated.")
	except Exception as e:
		return f"Error generating answer: {e}"

def load_text_file(filepath):
	"""Load and return the contents of a plain text file."""
	with open(filepath, 'r', encoding='utf-8') as f:
		return f.read()

def load_markdown_file(filepath):
	"""Load and return the contents of a Markdown (.md) file as text."""
	with open(filepath, 'r', encoding='utf-8') as f:
		return f.read()

def load_excel_file(filepath):
	"""Load and return the contents of an Excel (.xlsx) file as a string."""
	wb = load_workbook(filepath, data_only=True)
	text = []
	for sheet in wb.worksheets:
		text.append(f"Sheet: {sheet.title}")
		for row in sheet.iter_rows(values_only=True):
			text.append('\t'.join([str(cell) if cell is not None else '' for cell in row]))
	return '\n'.join(text)

def load_pdf_file(filepath):
	"""Extract and return text from a PDF file."""
	return extract_pdf_text(filepath)

def load_word_file(filepath):
	"""Extract and return text from a Word (.docx) file."""
	doc = Document(filepath)
	return '\n'.join([para.text for para in doc.paragraphs])

def load_document(filepath):
	"""Detect file type and load document accordingly."""
	ext = os.path.splitext(filepath)[1].lower()
	if ext == '.txt':
		return load_text_file(filepath)
	elif ext == '.pdf':
		return load_pdf_file(filepath)
	elif ext == '.docx':
		return load_word_file(filepath)
	elif ext == '.md':
		return load_markdown_file(filepath)
	elif ext == '.xlsx':
		return load_excel_file(filepath)
	else:
		raise ValueError(f"Unsupported file type: {ext}")

def chunk_text(text, chunk_size=1000, chunk_overlap=200):
	"""Split text into chunks using LangChain's RecursiveCharacterTextSplitter."""
	splitter = RecursiveCharacterTextSplitter(
		chunk_size=chunk_size,
		chunk_overlap=chunk_overlap
	)
	return splitter.split_text(text)
